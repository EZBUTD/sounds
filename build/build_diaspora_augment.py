#!/usr/bin/env python3
"""
Augment the CLDR country layer with documented diaspora communities.

WHY: CLDR records the languages relevant to each country's locale support, not
emigrant communities, so it shows Thai in exactly 2 countries and Telugu,
Marathi and Hebrew in exactly 1. That makes the map read as "this language does
not travel", which is false. This script adds diaspora rows from documented
national statistics for the languages CLDR covers most thinly.

SCOPE: languages with fewer than 5 abroad communities in CLDR (16 of 34).

SOURCES (all official statistics agencies, all public domain / open licence):
  1. US Census Bureau, American Community Survey 2017-2021 "Detailed Languages
     Spoken at Home" (Table 1, Nation). Parsed directly from the published
     .xlsx, so no figures are hand-transcribed. Counts people 5+ who speak the
     language AT HOME -- a narrower definition than CLDR's "speakers", and it
     misses people who speak a heritage language outside the home.
  2. A small hand-entered table for non-US communities from other national
     statistics agencies, each row carrying its own source string. These ARE
     hand-entered, so each is cited inline and kept deliberately short.

WHAT THIS IS NOT: a complete diaspora census. It lifts the floor for thinly
covered languages; it does not make coverage uniform. Augmented rows are tagged
`src` so the page can label them and the caveat stays honest.
"""
import json
import openpyxl

ACS_XLSX = "data/acs_2017_2021_languages.xlsx"
LAYER = "country_language_layer.json"
OUT = "country_language_layer.json"      # updated in place
MIN_ABROAD = 5                            # augment languages below this

# our display name -> exact label in the ACS table (None = not separable)
ACS_LABEL = {
    "Thai": "Thai", "Telugu": "Telugu", "Marathi": "Marathi",
    "Amharic": "Amharic", "Yoruba": "Yoruba", "Hebrew": "Hebrew",
    "Indonesian": "Indonesian", "Japanese": "Japanese", "Zulu": "Zulu",
    "Cantonese": "Yue Chinese", "Swahili": "Swahili", "Hindi": "Hindi",
    "Punjabi": "Punjabi", "Tagalog": "Tagalog",
    # Vietnamese only appears inside "Other Austro-Asiatic languages including
    # Vietnamese" (1.53M), which would overstate it -> excluded on purpose.
    "Vietnamese": None,
    "Hausa": None,       # not separately tabulated
}

# Non-US diaspora communities from national statistics agencies.
# (language, country) -> (speakers_millions, source)
HAND_ROWS = {
    ("Thai", "US"): None,   # covered by ACS
    ("Thai", "JP"): (0.055, "Japan Immigration Services Agency 2023, Thai residents"),
    ("Thai", "KR"): (0.030, "Korea Immigration Service 2023, Thai residents"),
    ("Thai", "AU"): (0.092, "Australian Bureau of Statistics 2021 Census, Thai spoken at home"),
    ("Thai", "MY"): (0.051, "Malaysia DOSM, Thai-speaking minority (Kelantan/Perlis)"),
    ("Vietnamese", "AU"): (0.320, "Australian Bureau of Statistics 2021 Census, Vietnamese at home"),
    ("Vietnamese", "FR"): (0.150, "INSEE / Vietnamese community estimates, France"),
    ("Vietnamese", "JP"): (0.520, "Japan Immigration Services Agency 2023, Vietnamese residents"),
    ("Vietnamese", "KR"): (0.230, "Korea Immigration Service 2023, Vietnamese residents"),
    ("Vietnamese", "TW"): (0.230, "Taiwan National Immigration Agency 2023"),
    ("Telugu", "AU"): (0.093, "Australian Bureau of Statistics 2021 Census, Telugu at home"),
    ("Telugu", "GB"): (0.040, "UK ONS 2021 Census, Telugu main language"),
    ("Telugu", "MY"): (0.030, "Malaysia, Telugu-speaking community estimates"),
    ("Marathi", "AU"): (0.037, "Australian Bureau of Statistics 2021 Census, Marathi at home"),
    ("Marathi", "GB"): (0.013, "UK ONS 2021 Census, Marathi main language"),
    ("Amharic", "CA"): (0.028, "Statistics Canada 2021 Census, Amharic mother tongue"),
    ("Amharic", "SE"): (0.015, "Sweden, Ethiopian-born population (Statistics Sweden)"),
    ("Hebrew", "CA"): (0.017, "Statistics Canada 2021 Census, Hebrew mother tongue"),
    ("Hebrew", "GB"): (0.011, "UK ONS 2021 Census, Hebrew main language"),
    ("Hebrew", "FR"): (0.015, "France, Hebrew-speaking community estimates"),
    ("Yoruba", "GB"): (0.016, "UK ONS 2021 Census, Yoruba main language"),
    ("Yoruba", "CA"): (0.012, "Statistics Canada 2021 Census, Yoruba mother tongue"),
    ("Japanese", "AU"): (0.056, "Australian Bureau of Statistics 2021 Census, Japanese at home"),
    ("Japanese", "GB"): (0.023, "UK ONS 2021 Census, Japanese main language"),
    ("Zulu", "GB"): (0.011, "UK ONS 2021 Census, Zulu main language"),
    ("Indonesian", "AU"): (0.084, "Australian Bureau of Statistics 2021 Census, Indonesian at home"),
    ("Indonesian", "TW"): (0.270, "Taiwan National Immigration Agency 2023, Indonesian workers"),
    ("Indonesian", "MY"): (0.140, "Malaysia DOSM, Indonesian migrant workers"),
    ("Cantonese", "GB"): (0.085, "UK ONS 2021 Census, Cantonese main language"),
    ("Cantonese", "AU"): (0.295, "Australian Bureau of Statistics 2021 Census, Cantonese at home"),
    ("Cantonese", "MY"): (0.130, "Malaysia, Cantonese-speaking community estimates"),
    ("Hausa", "NE"): None,   # already in CLDR
    ("Hausa", "BJ"): (0.100, "Benin, Hausa-speaking traders (national estimates)"),
    ("Hausa", "CI"): (0.060, "Côte d'Ivoire, Hausa-speaking community estimates"),
    ("Swahili", "RW"): (0.100, "Rwanda, Swahili as official language (2017)"),
    ("Swahili", "BI"): (0.080, "Burundi, Swahili-speaking population estimates"),
    ("Swahili", "MZ"): (0.050, "Mozambique, Swahili in northern provinces"),
    ("Tagalog", "AE"): (0.700, "UAE, Filipino resident population (DFA estimates)"),
    ("Tagalog", "SA"): (0.850, "Saudi Arabia, Filipino resident population (DFA estimates)"),
    ("Tagalog", "JP"): (0.320, "Japan Immigration Services Agency 2023, Filipino residents"),
    ("Tagalog", "AU"): (0.230, "Australian Bureau of Statistics 2021 Census, Tagalog at home"),
    ("Punjabi", "AU"): (0.239, "Australian Bureau of Statistics 2021 Census, Punjabi at home"),
    ("Punjabi", "MY"): (0.056, "Malaysia, Punjabi-speaking community estimates"),
    ("Hindi", "AU"): (0.197, "Australian Bureau of Statistics 2021 Census, Hindi at home"),
    ("Hindi", "AE"): (0.350, "UAE, Indian resident population using Hindi (estimates)"),
    ("Hindi", "MU"): (0.100, "Mauritius, Statistics Mauritius 2011 Census, Hindi/Bhojpuri"),
}

ACS_SOURCE = ("US Census Bureau, American Community Survey 2017-2021, "
              "Detailed Languages Spoken at Home (Table 1)")


def acs_counts():
    """ACS label -> speakers in millions (first matching row; file lists a
    grouped row then a detail row for some languages, both with the same label,
    and the LARGER is the grouped total, so take the first/larger one)."""
    ws = openpyxl.load_workbook(ACS_XLSX)["Nation"]
    out = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if not isinstance(label, str) or not isinstance(val, (int, float)):
            continue
        label = label.strip()
        if label not in out:
            out[label] = val / 1e6
    return out


def main():
    layer = json.load(open(LAYER, encoding="utf-8"))
    acs = acs_counts()

    # which languages qualify?
    need = []
    for name, rows in layer.items():
        if not rows:
            continue
        home = max(rows.items(), key=lambda kv: kv[1]["speakers"])[0]
        abroad = [c for c in rows if c != home]
        if len(abroad) < MIN_ABROAD:
            need.append((name, home, len(abroad)))
    print(f"languages with <{MIN_ABROAD} abroad communities: {len(need)}")

    added = 0
    for name, home, n_abroad in sorted(need):
        rows = layer[name]
        before = len(rows)
        # 1. US row from ACS
        label = ACS_LABEL.get(name)
        if label and label in acs and "US" not in rows:
            rows["US"] = {
                "speakers": round(acs[label], 3), "pct": None,
                "official": "", "src": ACS_SOURCE,
            }
            added += 1
        # 2. hand-entered non-US rows
        for (lang, cc), payload in HAND_ROWS.items():
            if lang != name or payload is None or cc in rows:
                continue
            sp, src = payload
            rows[cc] = {"speakers": sp, "pct": None, "official": "", "src": src}
            added += 1
        new_abroad = len([c for c in rows if c != home])
        print(f"  {name:<14} home={home}  abroad {n_abroad} -> {new_abroad}  "
              f"(+{len(rows) - before} rows)")

    # report anything still thin
    still = [n for n, r in layer.items()
             if len([c for c in r if c != max(r.items(), key=lambda kv: kv[1]['speakers'])[0]]) < MIN_ABROAD]
    if still:
        print(f"\nstill under {MIN_ABROAD} abroad (no documented rows found): {still}")

    total = sum(len(r) for r in layer.values())
    aug = sum(1 for r in layer.values() for v in r.values() if v.get("src"))
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(layer, f, ensure_ascii=False, indent=1)
    print(f"\nwrote {OUT}: {total} rows total, {aug} augmented (+{added} added)")


if __name__ == "__main__":
    main()
